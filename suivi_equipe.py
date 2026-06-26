"""
MODÈLE DE SUIVI D'ÉQUIPE – FOOTBALL JEUNES
Entente Sportive de Genech

USAGE (Mac) :
    pip3 install openpyxl pillow
    python3 suivi_equipe.py

→ Ouvre suivi_equipe.xlsx dans Numbers (double-clic)

LOGO : placez votre logo (PNG transparent) dans le même dossier
       sous le nom "logo_club.png"
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ─── Config globale ───────────────────────────────────────────────────────────

CLUB_NAME    = "Entente Sportive de Genech"
SAISON       = "2025/2026"
CATEGORIE    = "U13"          # ← à modifier selon l'équipe
LOGO_FILE    = "logo_club.png"

NB_PLAYERS   = 25
SESSIONS_PM  = 9              # créneaux par mois (colonnes séances)
NB_MATCHS_P3 = 30             # matchs convocations
NB_MATCHS_P4 = 20             # matchs feuille de match

MONTHS = ["SEPTEMBRE","OCTOBRE","NOVEMBRE","DÉCEMBRE",
          "JANVIER","FÉVRIER","MARS","AVRIL","MAI","JUIN"]

MONTH_COLORS = [
    "1565C0","6A1B9A","00695C","BF360C",
    "4527A0","AD1457","2E7D32","F57F17",
    "0277BD","558B2F"
]

# Couleurs principales
C_NAVY   = "0D1B3E"
C_BLUE   = "1565C0"
C_GOLD   = "C9A84C"
C_WHITE  = "FFFFFF"
C_LIGHT  = "F0F4FA"
C_STRIP  = "E8EDF6"
C_HDR    = "1A2E5A"
C_NAME   = "D6E4F0"
C_TOT    = "BBD6F5"
C_FOOTER = "263A5A"

# ─── Helpers styles ───────────────────────────────────────────────────────────

def F(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_NAVY, size=10, italic=False, name="Arial"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border(color=C_NAVY):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def col(n):
    return get_column_letter(n)

# ─── Cellule helper ───────────────────────────────────────────────────────────

def sc(ws, row, column, value=None, formula=None,
       bg=None, bold=False, color=C_NAVY, size=10,
       h="left", v="center", wrap=False,
       fmt=None, italic=False, brd=False, brd_color="CCCCCC"):
    c = ws.cell(row=row, column=column)
    c.value = formula if formula else value
    if bg:   c.fill = F(bg)
    c.font  = font(bold=bold, color=color, size=size, italic=italic)
    c.alignment = align(h, v, wrap)
    if fmt:  c.number_format = fmt
    if brd:  c.border = thin_border(brd_color)
    return c

def mc(ws, r1, c1, r2, c2, value=None, formula=None,
       bg=None, bold=False, color=C_NAVY, size=10,
       h="center", v="center", wrap=False,
       fmt=None, italic=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return sc(ws, r1, c1, value=value, formula=formula,
              bg=bg, bold=bold, color=color, size=size,
              h=h, v=v, wrap=wrap, fmt=fmt, italic=italic)

def row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def col_w(ws, column, width):
    ws.column_dimensions[col(column)].width = width

def add_logo(ws, path, anchor, size=(80, 80)):
    """Ajoute le logo si le fichier existe"""
    if not os.path.exists(path):
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(path)
        img.width  = size[0]
        img.height = size[1]
        img.anchor = anchor
        ws.add_image(img)
    except Exception:
        pass  # Pillow non installé ou erreur image

def divider_row(ws, row, num_cols, bg="E2E8F4"):
    """Ligne séparatrice visuelle"""
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=row, column=c_idx).fill = F(bg)
    row_h(ws, row, 6)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 1 : LISTE DES JOUEURS
# ─────────────────────────────────────────────────────────────────────────────

def build_page1(ws):
    ws.title = "📋 Liste Joueurs"
    ws.sheet_view.showGridLines = False

    col_w(ws, 1, 5)   # marge
    col_w(ws, 2, 8)   # N°
    col_w(ws, 3, 22)  # Nom
    col_w(ws, 4, 22)  # Prénom
    col_w(ws, 5, 16)  # Naissance
    col_w(ws, 6, 16)  # Poste
    col_w(ws, 7, 26)  # Contact
    col_w(ws, 8, 5)   # marge

    # ── Bandeau haut ─────────────────────────────────────────────────────────
    for r in range(1, 5):
        for c_idx in range(1, 9):
            ws.cell(row=r, column=c_idx).fill = F(C_NAVY)
        row_h(ws, r, 8)

    row_h(ws, 2, 55)
    mc(ws, 2,2,2,7, f"{CLUB_NAME}  –  {CATEGORIE}",
       bg=C_NAVY, bold=True, color=C_GOLD, size=16)

    # Logo
    add_logo(ws, LOGO_FILE, "B2", size=(72, 72))

    row_h(ws, 3, 20)
    mc(ws, 3,2,3,7, f"LISTE DES JOUEURS  –  Saison {SAISON}",
       bg=C_NAVY, bold=True, color=C_WHITE, size=11)

    divider_row(ws, 5, 8, "C9A84C")  # trait or

    # ── En-têtes colonnes ────────────────────────────────────────────────────
    row_h(ws, 6, 26)
    headers = ["", "N°", "Nom", "Prénom", "Date de naissance", "Poste", "Contact parent", ""]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=i)
        c.value = h
        c.fill  = F(C_HDR)
        c.font  = font(bold=True, color=C_WHITE, size=10)
        c.alignment = align("center", "center")
        c.border = thin_border(C_GOLD)

    # ── Lignes joueurs ───────────────────────────────────────────────────────
    for r in range(7, 7 + NB_PLAYERS):
        row_h(ws, r, 22)
        bg_row = C_STRIP if r % 2 == 0 else C_WHITE
        num = r - 6
        sc(ws, r, 1, bg=bg_row)
        sc(ws, r, 2, num, bg=bg_row, h="center", bold=True, color=C_BLUE, brd=True)
        for c_idx in range(3, 8):
            sc(ws, r, c_idx, bg=bg_row, brd=True)
        sc(ws, r, 8, bg=bg_row)

    # ── Pied de page ─────────────────────────────────────────────────────────
    foot = 7 + NB_PLAYERS
    divider_row(ws, foot, 8, C_GOLD)
    row_h(ws, foot + 1, 18)
    mc(ws, foot+1, 2, foot+1, 7, f"Saison {SAISON}  –  {CLUB_NAME}",
       bg=C_NAVY, italic=True, color=C_WHITE, size=9)

    ws.freeze_panes = "B7"

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 2 : PRÉSENCE ENTRAÎNEMENTS
# ─────────────────────────────────────────────────────────────────────────────

def build_page2(ws):
    ws.title = "📅 Présences Entraînement"
    ws.sheet_view.showGridLines = False

    COLS_PM       = SESSIONS_PM + 2      # séances + TOT + %
    FIRST_DATA    = 2
    PLAYER_ROW    = 6                    # première ligne joueur
    TOTAL_NB_ROW  = PLAYER_ROW + NB_PLAYERS + 1   # ligne nb présents
    TOTAL_PCT_ROW = TOTAL_NB_ROW + 1               # ligne % présence

    # Colonnes TOTAL ANNÉE (après les mois)
    total_cols_start = FIRST_DATA + len(MONTHS) * COLS_PM
    YEAR_NB_COL  = total_cols_start
    YEAR_PCT_COL = total_cols_start + 1

    total_width = 1 + len(MONTHS) * COLS_PM + 2   # +2 pour TOTAL ANNÉE

    # Largeurs
    col_w(ws, 1, 24)  # noms
    for m_idx in range(len(MONTHS)):
        base = FIRST_DATA + m_idx * COLS_PM
        for s in range(SESSIONS_PM):
            col_w(ws, base + s, 7)
        col_w(ws, base + SESSIONS_PM,     6)   # TOT
        col_w(ws, base + SESSIONS_PM + 1, 6)   # %
    col_w(ws, YEAR_NB_COL,  8)
    col_w(ws, YEAR_PCT_COL, 8)

    # ── Bandeau titre ────────────────────────────────────────────────────────
    row_h(ws, 1, 28)
    mc(ws, 1, 1, 1, total_width,
       f"📅  SUIVI DE PRÉSENCE AUX ENTRAÎNEMENTS  –  {CLUB_NAME}  –  Saison {SAISON}  –  {CATEGORIE}",
       bg=C_NAVY, bold=True, color=C_GOLD, size=12)

    # ── Légende ──────────────────────────────────────────────────────────────
    row_h(ws, 2, 20)
    legends = [
        (1,  "P  =  Présent",              "27AE60", C_WHITE),
        (3,  "AJ  =  Absence Justifiée",   "F1C40F", C_NAVY),
        (5,  "AI  =  Absence Injustifiée", "E74C3C", C_WHITE),
        (7,  "B  =  Blessé / Malade",      "8E44AD", C_WHITE),
    ]
    for col_start, txt, bg, fg in legends:
        mc(ws, 2, col_start, 2, col_start + 1, txt,
           bg=bg, bold=True, color=fg, size=9)

    divider_row(ws, 3, total_width, C_GOLD)

    # ── Ligne 4 : mois ───────────────────────────────────────────────────────
    row_h(ws, 4, 22)
    col_cursor = FIRST_DATA
    for m_idx, month in enumerate(MONTHS):
        mc(ws, 4, col_cursor, 4, col_cursor + COLS_PM - 1,
           month, bg=MONTH_COLORS[m_idx], bold=True, color=C_WHITE, size=10)
        col_cursor += COLS_PM

    mc(ws, 4, YEAR_NB_COL,  4, YEAR_PCT_COL,
       "TOTAL ANNÉE", bg=C_GOLD, bold=True, color=C_NAVY, size=10)

    # ── Ligne 5 : sous-en-têtes séances ──────────────────────────────────────
    row_h(ws, 5, 28)
    sc(ws, 5, 1, "NOM / PRÉNOM", bg=C_HDR, bold=True, color=C_WHITE, h="center", v="center")

    col_cursor = FIRST_DATA
    for m_idx in range(len(MONTHS)):
        bg_m = MONTH_COLORS[m_idx]
        for s in range(1, SESSIONS_PM + 1):
            sc(ws, 5, col_cursor, f"Séance {s}",
               bg="D6E8FF", bold=True, h="center", size=8, v="center", color=C_NAVY, brd=True)
            col_cursor += 1
        sc(ws, 5, col_cursor, "TOT",
           bg=C_TOT, bold=True, h="center", size=9, v="center", color=C_NAVY, brd=True)
        col_cursor += 1
        sc(ws, 5, col_cursor, "%",
           bg=C_TOT, bold=True, h="center", size=9, v="center", color=C_NAVY, brd=True)
        col_cursor += 1

    sc(ws, 5, YEAR_NB_COL,  "Nb total", bg=C_GOLD, bold=True, h="center", size=8, color=C_NAVY, brd=True)
    sc(ws, 5, YEAR_PCT_COL, "% annuel", bg=C_GOLD, bold=True, h="center", size=8, color=C_NAVY, brd=True)

    # ── Dropdown ─────────────────────────────────────────────────────────────
    dv = DataValidation(
        type="list",
        formula1='"P,AJ,AI,B"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Choisir : P, AJ, AI ou B"
    )
    # Appliquer sur toutes les colonnes séances
    ranges = []
    c2 = FIRST_DATA
    for _ in MONTHS:
        ranges.append(f"{col(c2)}{PLAYER_ROW}:{col(c2+SESSIONS_PM-1)}{PLAYER_ROW+NB_PLAYERS-1}")
        c2 += COLS_PM
    dv.sqref = " ".join(ranges)
    ws.add_data_validation(dv)

    # ── Formatage conditionnel par valeur (P/AJ/AI/B) ────────────────────────
    # Chaque règle s'applique sur toutes les plages de séances
    # La formule utilise la cellule en haut à gauche de chaque plage (référence relative)
    cf_values = [
        ("P",  "27AE60", "FFFFFF"),   # Présent    → vert
        ("AJ", "F1C40F", "5D4E00"),   # Abs. just. → jaune
        ("AI", "E74C3C", "FFFFFF"),   # Abs. injus → rouge
        ("B",  "8E44AD", "FFFFFF"),   # Blessé     → violet
    ]
    for rng in ranges:
        top_left = rng.split(":")[0]   # ex: "B6"
        for val, bg_hex, fg_hex in cf_values:
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'{top_left}="{val}"'],
                fill=F(bg_hex),
                font=Font(bold=True, color=fg_hex, name="Arial", size=10)
            ))

    # ── Lignes joueurs ────────────────────────────────────────────────────────
    for r in range(PLAYER_ROW, PLAYER_ROW + NB_PLAYERS):
        player_idx = r - PLAYER_ROW + 2
        bg_row = C_STRIP if r % 2 == 0 else C_WHITE
        row_h(ws, r, 20)

        sc(ws, r, 1,
           formula=f"='📋 Liste Joueurs'!C{player_idx+5}&\" \"&'📋 Liste Joueurs'!D{player_idx+5}",
           bg=C_NAME, bold=True, color=C_NAVY, h="left")

        col_cursor = FIRST_DATA
        tot_cells = []
        for m_idx in range(len(MONTHS)):
            session_start = col_cursor

            for s in range(SESSIONS_PM):
                c_cell = ws.cell(row=r, column=col_cursor)
                c_cell.fill = F(bg_row)
                c_cell.alignment = align("center")
                c_cell.border = thin_border()
                col_cursor += 1

            # TOT
            s_c = col(session_start)
            e_c = col(session_start + SESSIONS_PM - 1)
            tot_ref = f"{col(col_cursor)}{r}"
            sc(ws, r, col_cursor,
               formula=f'=COUNTIF({s_c}{r}:{e_c}{r},"P")',
               bg=C_TOT, bold=True, h="center", color=C_NAVY, brd=True)
            tot_cells.append(tot_ref)
            col_cursor += 1

            # %
            cnt_rng = f"{col(session_start)}{r}:{col(session_start+SESSIONS_PM-1)}{r}"
            sc(ws, r, col_cursor,
               formula=f'=IFERROR({tot_ref}/COUNTA({cnt_rng}),"")',
               bg=C_TOT, bold=True, h="center", color=C_NAVY, fmt="0%", brd=True)
            col_cursor += 1

        # TOTAL ANNÉE – nb séances présent
        sc(ws, r, YEAR_NB_COL,
           formula="=SUM(" + "+".join(tot_cells) + ")",
           bg="FFF5CC", bold=True, h="center", color=C_NAVY, brd=True)

        # TOTAL ANNÉE – % (présences / total séances saisies)
        all_sessions_ref = []
        c3 = FIRST_DATA
        for _ in MONTHS:
            all_sessions_ref.append(f"{col(c3)}{r}:{col(c3+SESSIONS_PM-1)}{r}")
            c3 += COLS_PM
        all_rng = ",".join(all_sessions_ref)
        sc(ws, r, YEAR_PCT_COL,
           formula=f'=IFERROR({col(YEAR_NB_COL)}{r}/COUNTA({all_rng}),"")',
           bg="FFF5CC", bold=True, h="center", color=C_NAVY, fmt="0%", brd=True)

    # ── Ligne séparatrice ─────────────────────────────────────────────────────
    divider_row(ws, PLAYER_ROW + NB_PLAYERS, total_width, "C9D8EE")

    # ── Totaux bas : Nb présents ──────────────────────────────────────────────
    row_h(ws, TOTAL_NB_ROW, 22)
    sc(ws, TOTAL_NB_ROW, 1, "Nb présents / séance",
       bg=C_FOOTER, bold=True, color=C_WHITE, h="center", size=9)

    row_h(ws, TOTAL_PCT_ROW, 22)
    sc(ws, TOTAL_PCT_ROW, 1, "% présence / séance",
       bg=C_FOOTER, bold=True, color=C_GOLD, h="center", size=9)

    col_cursor = FIRST_DATA
    for m_idx in range(len(MONTHS)):
        bg_m_light = "EAF3FF"
        for s in range(SESSIONS_PM):
            c_letter = col(col_cursor)
            # Nb présents
            sc(ws, TOTAL_NB_ROW, col_cursor,
               formula=f'=COUNTIF({c_letter}{PLAYER_ROW}:{c_letter}{PLAYER_ROW+NB_PLAYERS-1},"P")',
               bg=C_FOOTER, bold=True, color=C_WHITE, h="center", size=9, brd=True, brd_color=C_NAVY)
            # % présence
            sc(ws, TOTAL_PCT_ROW, col_cursor,
               formula=f'=IFERROR({c_letter}{TOTAL_NB_ROW}/COUNTA({c_letter}{PLAYER_ROW}:{c_letter}{PLAYER_ROW+NB_PLAYERS-1}),"")',
               bg=C_FOOTER, bold=True, color=C_GOLD, h="center", fmt="0%", size=9, brd=True, brd_color=C_NAVY)
            col_cursor += 1

        # TOT mois
        sc(ws, TOTAL_NB_ROW, col_cursor, bg=C_FOOTER, bold=True, color=C_WHITE, h="center")
        sc(ws, TOTAL_PCT_ROW, col_cursor, bg=C_FOOTER, bold=True, color=C_GOLD, h="center")
        col_cursor += 1
        # % mois
        sc(ws, TOTAL_NB_ROW, col_cursor, bg=C_FOOTER, bold=True, color=C_WHITE, h="center")
        sc(ws, TOTAL_PCT_ROW, col_cursor, bg=C_FOOTER, bold=True, color=C_GOLD, h="center")
        col_cursor += 1

    # Totaux ANNÉE
    yr_col = col(YEAR_NB_COL)
    yp_col = col(YEAR_PCT_COL)
    sc(ws, TOTAL_NB_ROW, YEAR_NB_COL,
       formula=f"=SUM({yr_col}{PLAYER_ROW}:{yr_col}{PLAYER_ROW+NB_PLAYERS-1})",
       bg=C_GOLD, bold=True, color=C_NAVY, h="center", size=10, brd=True)
    sc(ws, TOTAL_PCT_ROW, YEAR_PCT_COL,
       formula=f'=IFERROR(AVERAGE({yp_col}{PLAYER_ROW}:{yp_col}{PLAYER_ROW+NB_PLAYERS-1}),"")',
       bg=C_GOLD, bold=True, color=C_NAVY, h="center", fmt="0%", size=10, brd=True)

    ws.freeze_panes = f"B{PLAYER_ROW}"

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 3 : CONVOCATIONS MATCHS
# ─────────────────────────────────────────────────────────────────────────────

def build_page3(ws):
    ws.title = "📣 Convocations Matchs"
    ws.sheet_view.showGridLines = False

    PLAYER_ROW = 6

    col_w(ws, 1, 24)
    for m in range(1, NB_MATCHS_P3 + 1):
        col_w(ws, m + 1, 12)

    total_cols = NB_MATCHS_P3 + 2

    # ── Titre ────────────────────────────────────────────────────────────────
    row_h(ws, 1, 28)
    mc(ws, 1,1,1,total_cols,
       f"📣  CONVOCATIONS MATCHS  –  {CLUB_NAME}  –  {CATEGORIE}  –  Saison {SAISON}",
       bg=C_NAVY, bold=True, color=C_GOLD, size=12)

    # ── Légende ──────────────────────────────────────────────────────────────
    row_h(ws, 2, 20)
    leg = [
        (1,  2,  "✅  Convoqué",        "27AE60", C_WHITE),
        (3,  4,  "❌  Non convoqué",    "E74C3C", C_WHITE),
        (5,  7,  "⚠️  Convoqué absent", "F39C12", C_WHITE),
    ]
    for c1, c2, txt, bg, fg in leg:
        mc(ws, 2,c1,2,c2, txt, bg=bg, bold=True, color=fg, size=9)

    divider_row(ws, 3, total_cols, C_GOLD)
    row_h(ws, 4, 14)  # espace

    # ── En-têtes matchs ──────────────────────────────────────────────────────
    row_h(ws, 5, 24)
    sc(ws, 5, 1, "NOM / PRÉNOM", bg=C_HDR, bold=True, color=C_WHITE, h="center", brd=True)

    MATCH_COLORS = ["1A5276","1B4F72","154360","0E6655","145A32",
                    "4A235A","6E2F7A","7B241C","78281F","1F618D"]
    for m in range(1, NB_MATCHS_P3 + 1):
        bg_m = MATCH_COLORS[(m-1) % len(MATCH_COLORS)]
        sc(ws, 5, m+1, f"Match {m}", bg=bg_m, bold=True, color=C_WHITE,
           h="center", size=9, brd=True)

    # ── Dropdown convocation ──────────────────────────────────────────────────
    dv = DataValidation(
        type="list",
        formula1='"✅ Convoqué,❌ Non convoqué,⚠️ Convoqué absent"',
        showDropDown=False
    )
    dv.sqref = f"B{PLAYER_ROW}:{col(NB_MATCHS_P3+1)}{PLAYER_ROW+NB_PLAYERS-1}"
    ws.add_data_validation(dv)

    # ── Lignes joueurs ────────────────────────────────────────────────────────
    for r in range(PLAYER_ROW, PLAYER_ROW + NB_PLAYERS):
        player_idx = r - PLAYER_ROW + 7
        bg_row = C_STRIP if r % 2 == 0 else C_WHITE
        row_h(ws, r, 20)
        sc(ws, r, 1,
           formula=f"='📋 Liste Joueurs'!C{player_idx}&\" \"&'📋 Liste Joueurs'!D{player_idx}",
           bg=C_NAME, bold=True, color=C_NAVY)
        for m in range(1, NB_MATCHS_P3 + 1):
            c = ws.cell(row=r, column=m+1)
            c.fill = F(bg_row)
            c.alignment = align("center")
            c.border = thin_border()

    # ── Totaux bas ────────────────────────────────────────────────────────────
    divider_row(ws, PLAYER_ROW + NB_PLAYERS, total_cols, "C9D8EE")
    tot_row = PLAYER_ROW + NB_PLAYERS + 1
    pct_row = tot_row + 1
    row_h(ws, tot_row, 20)
    row_h(ws, pct_row, 20)

    sc(ws, tot_row, 1, "Nb convoqués / match", bg=C_FOOTER, bold=True, color=C_WHITE, h="center", size=9)
    sc(ws, pct_row, 1, "% convocation / match", bg=C_FOOTER, bold=True, color=C_GOLD,  h="center", size=9)

    for m in range(1, NB_MATCHS_P3 + 1):
        c_letter = col(m+1)
        sc(ws, tot_row, m+1,
           formula=f'=COUNTIF({c_letter}{PLAYER_ROW}:{c_letter}{PLAYER_ROW+NB_PLAYERS-1},"✅ Convoqué")',
           bg=C_FOOTER, bold=True, color=C_WHITE, h="center", size=9, brd=True, brd_color=C_NAVY)
        sc(ws, pct_row, m+1,
           formula=f'=IFERROR({c_letter}{tot_row}/COUNTA({c_letter}{PLAYER_ROW}:{c_letter}{PLAYER_ROW+NB_PLAYERS-1}),"")',
           bg=C_FOOTER, bold=True, color=C_GOLD, h="center", fmt="0%", size=9, brd=True, brd_color=C_NAVY)

    ws.freeze_panes = f"B{PLAYER_ROW}"

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 4 : FEUILLE DE MATCH
# ─────────────────────────────────────────────────────────────────────────────

def build_page4(ws):
    ws.title = "⚽ Feuille de Match"
    ws.sheet_view.showGridLines = False

    COLS_PM4    = 4
    PLAYER_ROW  = 9
    FIRST_MATCH = 2

    col_w(ws, 1, 24)
    for m in range(NB_MATCHS_P4):
        base = FIRST_MATCH + m * COLS_PM4
        col_w(ws, base,   11)   # Tit/Rem
        col_w(ws, base+1,  8)   # Min
        col_w(ws, base+2,  7)   # Buts
        col_w(ws, base+3,  8)   # PD

    total_cols = 1 + NB_MATCHS_P4 * COLS_PM4

    # ── Titre ────────────────────────────────────────────────────────────────
    row_h(ws, 1, 28)
    mc(ws, 1,1,1,total_cols,
       f"⚽  FEUILLE DE MATCH  –  {CLUB_NAME}  –  {CATEGORIE}  –  Saison {SAISON}",
       bg="1B4F1F", bold=True, color=C_GOLD, size=12)

    divider_row(ws, 2, total_cols, C_GOLD)

    # Labels colonne A
    row_labels = {3:"MATCH", 4:"Date", 5:"Type", 6:"Dom / Ext", 7:"Score", 8:"NOM / PRÉNOM"}
    row_heights = {3:22, 4:22, 5:22, 6:22, 7:22, 8:26}
    for row_n, lbl in row_labels.items():
        row_h(ws, row_n, row_heights[row_n])
        sc(ws, row_n, 1, lbl, bg=C_HDR, bold=True, color=C_WHITE,
           h="center", v="center", size=9 if row_n == 8 else 10)

    # Dropdowns
    dv_type   = DataValidation(type="list", formula1='"Championnat,Amical,Coupe,Tournoi"', showDropDown=False)
    dv_domext = DataValidation(type="list", formula1='"Domicile,Extérieur,Terrain neutre"', showDropDown=False)
    dv_role   = DataValidation(type="list", formula1='"Titulaire,Remplaçant,N/A"',          showDropDown=False)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_domext)
    ws.add_data_validation(dv_role)

    MATCH_HDR_COLORS = [
        "1A5276","1B4F72","154360","0E6655","145A32",
        "4A235A","6E2F7A","7B241C","78281F","1F618D",
        "17202A","1B2631","0B5345","4D5656","4A4A4A",
        "2C3E50","641E16","1A5276","196F3D","7D6608"
    ]

    type_cells   = []
    domext_cells = []
    role_ranges  = []

    for m in range(NB_MATCHS_P4):
        start_col = FIRST_MATCH + m * COLS_PM4
        hdr_bg = MATCH_HDR_COLORS[m % len(MATCH_HDR_COLORS)]

        # Ligne 3 : Match N (merge 4 cols)
        mc(ws, 3, start_col, 3, start_col+COLS_PM4-1,
           f"Match {m+1}", bg=hdr_bg, bold=True, color=C_WHITE, size=10)

        # Ligne 4 : Date
        mc(ws, 4, start_col, 4, start_col+COLS_PM4-1,
           bg="EBF5FB", fmt="DD/MM/YYYY")
        ws.cell(row=4, column=start_col).alignment = align("center")

        # Ligne 5 : Type
        mc(ws, 5, start_col, 5, start_col+COLS_PM4-1, bg="E8F8F5")
        ws.cell(row=5, column=start_col).alignment = align("center")
        type_cells.append(f"{col(start_col)}5")

        # Ligne 6 : Dom/Ext
        mc(ws, 6, start_col, 6, start_col+COLS_PM4-1, bg="FEF9E7")
        ws.cell(row=6, column=start_col).alignment = align("center")
        domext_cells.append(f"{col(start_col)}6")

        # Ligne 7 : Score
        mc(ws, 7, start_col, 7, start_col+COLS_PM4-1, bg="FDEDEC")
        ws.cell(row=7, column=start_col).alignment = align("center")
        ws.cell(row=7, column=start_col).font = font(bold=True, size=11, color="C0392B")

        # Ligne 8 : Sous-en-têtes
        sub = ["Tit / Rem", "Min", "Buts", "Passes"]
        sub_bgs = ["A9DFBF","A9CCE3","F9E79F","D2B4DE"]
        for sc_idx in range(COLS_PM4):
            sc(ws, 8, start_col+sc_idx, sub[sc_idx],
               bg=sub_bgs[sc_idx], bold=True, h="center", size=8, v="center",
               color=C_NAVY, brd=True)

        # Lignes joueurs
        for r in range(PLAYER_ROW, PLAYER_ROW + NB_PLAYERS):
            bg_row = C_STRIP if r % 2 == 0 else C_WHITE
            for sc_idx in range(COLS_PM4):
                c = ws.cell(row=r, column=start_col+sc_idx)
                c.fill = F(bg_row)
                c.alignment = align("center")
                c.border = thin_border()

        role_ranges.append(f"{col(start_col)}{PLAYER_ROW}:{col(start_col)}{PLAYER_ROW+NB_PLAYERS-1}")

    # Appliquer dropdowns
    dv_type.sqref   = " ".join(type_cells)
    dv_domext.sqref = " ".join(domext_cells)
    dv_role.sqref   = " ".join(role_ranges)

    # Noms joueurs
    for r in range(PLAYER_ROW, PLAYER_ROW + NB_PLAYERS):
        player_idx = r - PLAYER_ROW + 7
        row_h(ws, r, 20)
        sc(ws, r, 1,
           formula=f"='📋 Liste Joueurs'!C{player_idx}&\" \"&'📋 Liste Joueurs'!D{player_idx}",
           bg=C_NAME, bold=True, color=C_NAVY)

    ws.freeze_panes = f"B{PLAYER_ROW}"

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 5 : STATS GÉNÉRALES
# ─────────────────────────────────────────────────────────────────────────────

def build_page5(ws):
    ws.title = "📊 Stats Générales"
    ws.sheet_view.showGridLines = False

    PLAYER_ROW  = 6
    COLS_PM4    = 4
    COLS_PM_P2  = SESSIONS_PM + 2
    FIRST_DATA_P2  = 2
    FIRST_MATCH_P3 = 2
    FIRST_MATCH_P4 = 2

    col_widths = [24, 13, 13, 13, 13, 14, 12, 10, 13, 16]
    for i, w in enumerate(col_widths, 1):
        col_w(ws, i, w)

    headers = [
        "NOM / PRÉNOM",
        "Taux présence\nentraîn.",
        "Matchs\nconvoqué",
        "Taux\nconvocation",
        "Matchs\ntitulaire",
        "Matchs\nremplaçant",
        "Minutes\njouées",
        "Buts",
        "Passes\ndécisives",
        "Indice IRE\n(Conv ÷ Présence)"
    ]

    total_cols = len(headers)

    # ── Titre ────────────────────────────────────────────────────────────────
    row_h(ws, 1, 28)
    mc(ws, 1,1,1,total_cols,
       f"📊  STATISTIQUES GÉNÉRALES  –  {CLUB_NAME}  –  {CATEGORIE}  –  Saison {SAISON}",
       bg=C_NAVY, bold=True, color=C_GOLD, size=12)

    # ── Explication IRE ──────────────────────────────────────────────────────
    row_h(ws, 2, 20)
    mc(ws, 2,1,2,total_cols,
       "Indice IRE :   🟢 1,00 – 1,10 = Justement convoqué     🟠 < 1,00 = Sous-convoqué     🔴 > 1,10 = Sur-convoqué",
       bg="EEF2FA", italic=True, color=C_HDR, size=9)

    divider_row(ws, 3, total_cols, C_GOLD)
    row_h(ws, 4, 14)

    # ── En-têtes ─────────────────────────────────────────────────────────────
    row_h(ws, 5, 46)
    for i, h in enumerate(headers, 1):
        sc(ws, 5, i, h, bg=C_HDR, bold=True, color=C_WHITE,
           h="center", v="center", wrap=True, size=9, brd=True)

    # ── Lignes joueurs ────────────────────────────────────────────────────────
    for r in range(PLAYER_ROW, PLAYER_ROW + NB_PLAYERS):
        player_idx = r - PLAYER_ROW + 7
        bg_row = C_STRIP if r % 2 == 0 else C_WHITE
        row_h(ws, r, 22)

        # Col A : Nom
        sc(ws, r, 1,
           formula=f"='📋 Liste Joueurs'!C{player_idx}&\" \"&'📋 Liste Joueurs'!D{player_idx}",
           bg=C_NAME, bold=True, color=C_NAVY)

        # Page 2 : ligne joueurs commence à ligne 6, même index
        p2_row = r  # même numéro de ligne

        # Colonnes % de chaque mois (col FIRST_DATA + m*COLS_PM + SESSIONS_PM + 1)
        pct_cols_p2 = [FIRST_DATA_P2 + m * COLS_PM_P2 + SESSIONS_PM + 1 for m in range(len(MONTHS))]
        pct_refs = ",".join([f"'📅 Présences Entraînement'!{col(c)}{p2_row}" for c in pct_cols_p2])
        sc(ws, r, 2,
           formula=f"=IFERROR(AVERAGEIF({{{pct_refs}}},\"<>\"),\"\")",
           bg=bg_row, h="center", fmt="0%", brd=True)

        # Page 3 : ligne joueurs commence à ligne 6
        p3_row = r
        sc(ws, r, 3,
           formula=f"=COUNTIF('📣 Convocations Matchs'!B{p3_row}:{col(NB_MATCHS_P3+1)}{p3_row},\"✅ Convoqué\")",
           bg=bg_row, h="center", brd=True)

        sc(ws, r, 4,
           formula=f"=IFERROR(C{r}/COUNTA('📣 Convocations Matchs'!B{p3_row}:{col(NB_MATCHS_P3+1)}{p3_row}),\"\")",
           bg=bg_row, h="center", fmt="0%", brd=True)

        # Page 4 : ligne joueurs commence à ligne 9
        p4_row = r + 3  # PLAYER_ROW_P4 = 9, PLAYER_ROW_P5 = 6 → offset = 3

        tit = ",".join([f"'⚽ Feuille de Match'!{col(FIRST_MATCH_P4+m*COLS_PM4)}{p4_row}" for m in range(NB_MATCHS_P4)])
        min_ = ",".join([f"'⚽ Feuille de Match'!{col(FIRST_MATCH_P4+m*COLS_PM4+1)}{p4_row}" for m in range(NB_MATCHS_P4)])
        but  = ",".join([f"'⚽ Feuille de Match'!{col(FIRST_MATCH_P4+m*COLS_PM4+2)}{p4_row}" for m in range(NB_MATCHS_P4)])
        pd_  = ",".join([f"'⚽ Feuille de Match'!{col(FIRST_MATCH_P4+m*COLS_PM4+3)}{p4_row}" for m in range(NB_MATCHS_P4)])

        sc(ws, r, 5, formula=f'=COUNTIF({{{tit}}},"Titulaire")',  bg=bg_row, h="center", brd=True)
        sc(ws, r, 6, formula=f'=COUNTIF({{{tit}}},"Remplaçant")', bg=bg_row, h="center", brd=True)
        sc(ws, r, 7, formula=f"=SUM({{{min_}}})", bg=bg_row, h="center", brd=True)
        sc(ws, r, 8, formula=f"=SUM({{{but}}})",  bg=bg_row, h="center", brd=True)
        sc(ws, r, 9, formula=f"=SUM({{{pd_}}})",  bg=bg_row, h="center", brd=True)

        sc(ws, r, 10,
           formula=f"=IFERROR(D{r}/B{r},\"\")",
           bg=bg_row, bold=True, h="center", fmt="0.00", brd=True)

    # ── Formatage conditionnel IRE ────────────────────────────────────────────
    ire_range = f"J{PLAYER_ROW}:J{PLAYER_ROW+NB_PLAYERS-1}"

    ws.conditional_formatting.add(ire_range, FormulaRule(
        formula=[f"AND(J{PLAYER_ROW}<>\"\",J{PLAYER_ROW}>1.1)"],
        fill=F("FFCDD2"), font=Font(color="C0392B", bold=True, name="Arial")))

    ws.conditional_formatting.add(ire_range, FormulaRule(
        formula=[f"AND(J{PLAYER_ROW}>=1,J{PLAYER_ROW}<=1.1)"],
        fill=F("D5F5E3"), font=Font(color="1E8449", bold=True, name="Arial")))

    ws.conditional_formatting.add(ire_range, FormulaRule(
        formula=[f"AND(J{PLAYER_ROW}<>\"\",J{PLAYER_ROW}<1)"],
        fill=F("FAD7A0"), font=Font(color="D35400", bold=True, name="Arial")))

    # ── Ligne moyenne équipe ──────────────────────────────────────────────────
    divider_row(ws, PLAYER_ROW + NB_PLAYERS, total_cols, "C9D8EE")
    avg_row = PLAYER_ROW + NB_PLAYERS + 1
    row_h(ws, avg_row, 24)
    sc(ws, avg_row, 1, "MOYENNE ÉQUIPE", bg=C_FOOTER, bold=True, color=C_GOLD, h="center", size=10)

    fmts = ["0%","0","0%","0","0","0","0","0","0.00"]
    for i, fmt in enumerate(fmts, 2):
        c_l = col(i)
        sc(ws, avg_row, i,
           formula=f"=IFERROR(AVERAGE({c_l}{PLAYER_ROW}:{c_l}{PLAYER_ROW+NB_PLAYERS-1}),\"\")",
           bg=C_FOOTER, bold=True, color=C_WHITE, h="center", fmt=fmt, brd=True, brd_color=C_NAVY)

    ws.freeze_panes = f"B{PLAYER_ROW}"

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("📋 Liste Joueurs",          build_page1),
        ("📅 Présences Entraînement", build_page2),
        ("📣 Convocations Matchs",    build_page3),
        ("⚽ Feuille de Match",       build_page4),
        ("📊 Stats Générales",        build_page5),
    ]

    for name, builder in sheets:
        ws = wb.create_sheet(name)
        print(f"  {name}…")
        builder(ws)

    output = "suivi_equipe.xlsx"
    wb.save(output)

    print(f"\n✅  Fichier créé : {output}")
    if not os.path.exists(LOGO_FILE):
        print(f"⚠️   Logo introuvable — placez '{LOGO_FILE}' dans le même dossier et relancez.")
    print("\n→  Ouvrir avec Numbers : double-clic sur suivi_equipe.xlsx")
    print(f"\nCONSEIL : Remplissez d'abord '📋 Liste Joueurs' (club : {CLUB_NAME}, catégorie : {CATEGORIE})")
    print("          Modifiez CLUB_NAME et CATEGORIE en haut du script si besoin.")

if __name__ == "__main__":
    main()
